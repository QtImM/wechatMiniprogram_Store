from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "商品组编码",
    "商品名称",
    "商品类型",
    "分类ID",
    "分类名称",
    "关键词",
    "简介",
    "主图URL",
    "轮播图URL(多个用|分隔)",
    "详情说明",
    "详情图URL(多个用|分隔)",
    "SKU编码",
    "规格名称",
    "规格值",
    "售价(元)",
    "市场价(元)",
    "库存",
    "上架状态(上架/下架)",
    "排序",
    "SKU图片URL",
    "创建时间",
]

TYPE_OPTIONS = ["实物", "虚拟"]
STATUS_OPTIONS = ["上架", "下架"]
SPEC_NAME_OPTIONS = ["规格", "净含量", "包装", "口味", "套餐"]
CATEGORY_OPTIONS = [
    "滋补养生",
    "茶饮花茶",
    "零食坚果",
    "保健食品",
    "药膳食材",
    "花果茶",
    "胶类滋补",
    "轻养零食",
]

INSTRUCTIONS = [
    "1. 客户请始终基于本模板填写，不要修改表头名称、顺序和工作表名称。",
    "2. 同一商品如果有多个规格，请保持“商品组编码”一致；不同商品必须使用不同的商品组编码。",
    "3. 首次铺货请选择后台“新增商品”；后续改价、改库存、上下架，请先从后台导出现有商品后再修改并使用“更新已有商品”。",
    "4. 更新已有商品时，必须保留该商品全部 SKU 行，避免漏一行导致规格被误删。",
    "5. 价格单位为元，库存必须为非负整数，排序数值越大越靠前。",
    "6. 图片地址请使用客户正式图片地址；多张图片请使用英文竖线 | 分隔。",
    "7. 本模板中的示例文件分为“可直接参考上新”和“仅作格式参考”两类，请按文件名说明使用。",
]

README_TEXT = """药食同源 v1.0 商品 Excel 模板包

一、文件用途
1. 01-商品导入空白模板.xlsx
   直接给客户填写的正式空白模板。
2. 02-新增商品示例-单规格.xlsx
   演示一个商品只有一个 SKU 时如何填写，可直接照着改。
3. 03-新增商品示例-多规格.xlsx
   演示一个商品有多个 SKU 时如何填写，重点看“商品组编码”保持一致。
4. 04-更新商品示例-改价改库存.xlsx
   仅作格式参考。真实更新时必须先从后台导出目标商品，再在导出文件上修改。
5. 05-补充规格示例-新增SKU.xlsx
   仅作格式参考。真实补规格时建议先导出目标商品，再补新 SKU 行后导入。

二、客户最常用的 3 种操作
1. 首次上新
   用 01 空白模板填写，后台选择“新增商品”导入。
2. 改价格 / 改库存 / 改上下架
   先在后台导出商品，再按 04 示例格式修改，后台选择“更新已有商品”导入。
3. 给老商品补新规格
   先导出商品，再按 05 示例格式补行，后台选择“新增并补充规格”导入。

三、填写规则
1. 商品名称、商品类型、分类名称、主图URL、SKU编码、售价、库存、上架状态必填。
2. 同一商品的多行规格必须使用同一个“商品组编码”。
3. 建议 SKU 编码使用英文大写、数字和连字符，例如 GJHC-120G。
4. 商品详情请写客户可读的普通文案，不需要写 HTML。
5. 如果不确定怎么改，请先复制一个示例文件，再替换成客户自己的商品资料。

四、交付建议
1. 把本模板包连同后台操作说明一起发给客户。
2. 第一次导入建议先用 1 到 3 个商品试跑，确认格式无误后再批量导入。
3. 真正更新已有商品时，不要直接改示例文件，请先导出系统里的真实商品文件再改。
"""


def build_workbook(title: str, rows: list[list[str]]) -> Workbook:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "商品导入"
    dict_sheet = workbook.create_sheet("字典")
    instruction_sheet = workbook.create_sheet("填写说明")

    create_header(data_sheet)
    populate_rows(data_sheet, rows)
    populate_dictionary(dict_sheet)
    populate_instructions(instruction_sheet, title)
    apply_validations(data_sheet)
    apply_layout(data_sheet, dict_sheet, instruction_sheet)
    workbook.active = 0
    return workbook


def create_header(sheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="5F8F74")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
    sheet.freeze_panes = "A2"


def populate_rows(sheet, rows: list[list[str]]) -> None:
    for row_index, row_values in enumerate(rows, start=2):
        for column, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=column, value=value)


def populate_dictionary(sheet) -> None:
    sheet["A1"] = "分类名称"
    for index, value in enumerate(CATEGORY_OPTIONS, start=2):
        sheet[f"A{index}"] = value

    sheet["C1"] = "商品类型"
    for index, value in enumerate(TYPE_OPTIONS, start=2):
        sheet[f"C{index}"] = value

    sheet["E1"] = "上架状态"
    for index, value in enumerate(STATUS_OPTIONS, start=2):
        sheet[f"E{index}"] = value

    sheet["G1"] = "规格名称"
    for index, value in enumerate(SPEC_NAME_OPTIONS, start=2):
        sheet[f"G{index}"] = value


def populate_instructions(sheet, title: str) -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14)
    for row_index, line in enumerate(INSTRUCTIONS, start=3):
        sheet.cell(row=row_index, column=1, value=line)


def apply_validations(sheet) -> None:
    add_validation(sheet, "C2:C1001", TYPE_OPTIONS)
    add_validation(sheet, "E2:E1001", CATEGORY_OPTIONS)
    add_validation(sheet, "M2:M1001", SPEC_NAME_OPTIONS)
    add_validation(sheet, "R2:R1001", STATUS_OPTIONS)


def add_validation(sheet, cell_range: str, values: list[str]) -> None:
    escaped = ",".join(value.replace('"', '""') for value in values)
    validation = DataValidation(type="list", formula1=f'"{escaped}"', allow_blank=True)
    validation.prompt = "请从下拉选项中选择，或参照示例填写。"
    validation.error = "请输入模板允许的选项值。"
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def apply_layout(data_sheet, dict_sheet, instruction_sheet) -> None:
    widths = {
        "A": 16,
        "B": 22,
        "C": 12,
        "D": 12,
        "E": 16,
        "F": 20,
        "G": 26,
        "H": 26,
        "I": 34,
        "J": 30,
        "K": 34,
        "L": 18,
        "M": 12,
        "N": 14,
        "O": 12,
        "P": 12,
        "Q": 10,
        "R": 16,
        "S": 10,
        "T": 26,
        "U": 20,
    }
    for column, width in widths.items():
        data_sheet.column_dimensions[column].width = width
    dict_sheet.column_dimensions["A"].width = 18
    dict_sheet.column_dimensions["C"].width = 12
    dict_sheet.column_dimensions["E"].width = 14
    dict_sheet.column_dimensions["G"].width = 14
    instruction_sheet.column_dimensions["A"].width = 120
    instruction_sheet.freeze_panes = "A3"


def single_spec_rows() -> list[list[str]]:
    return [[
        "GJHC-001",
        "桂圆红枣姜茶",
        "实物",
        "",
        "茶饮花茶",
        "桂圆 红枣 姜茶 暖饮",
        "独立袋泡装，适合日常冲泡",
        "https://images.unsplash.com/photo-1596701062351-8c2c14d1fdd0?w=600",
        "https://images.unsplash.com/photo-1596701062351-8c2c14d1fdd0?w=600|https://images.unsplash.com/photo-1597481499750-3e6b22637e12?w=600",
        "桂圆、红枣与姜片搭配，适合秋冬热饮场景。",
        "https://images.unsplash.com/photo-1597481499750-3e6b22637e12?w=900",
        "GJHC-120G",
        "规格",
        "12袋",
        "39.90",
        "59.90",
        "200",
        "上架",
        "100",
        "https://images.unsplash.com/photo-1596701062351-8c2c14d1fdd0?w=600",
        "",
    ]]


def multi_spec_rows() -> list[list[str]]:
    common = [
        "RGEJ-001",
        "玫瑰阿胶糕礼盒",
        "实物",
        "",
        "胶类滋补",
        "阿胶 玫瑰 礼盒 滋补",
        "礼盒装阿胶糕，适合日常滋补和送礼",
        "https://images.unsplash.com/photo-1505252585461-04db1eb84625?w=600",
        "https://images.unsplash.com/photo-1505252585461-04db1eb84625?w=600|https://images.unsplash.com/photo-1587049352846-4a222e784d38?w=600",
        "同一商品多个净含量时，请保持商品组编码一致，每个规格单独一行。",
        "https://images.unsplash.com/photo-1505252585461-04db1eb84625?w=900|https://images.unsplash.com/photo-1587049352846-4a222e784d38?w=900",
    ]
    rows = []
    for sku_code, spec_value, price, market_price, stock in [
        ("RGEJ-120G", "120g", "69.00", "89.00", "80"),
        ("RGEJ-250G", "250g", "128.00", "168.00", "50"),
        ("RGEJ-500G", "500g", "228.00", "288.00", "30"),
    ]:
        rows.append(common + [
            sku_code,
            "净含量",
            spec_value,
            price,
            market_price,
            stock,
            "上架",
            "120",
            "https://images.unsplash.com/photo-1505252585461-04db1eb84625?w=600",
            "",
        ])
    return rows


def update_rows() -> list[list[str]]:
    common = [
        "SPU-10001",
        "示例：导出后修改的老商品",
        "实物",
        "",
        "滋补养生",
        "示例 更新 改价 改库存",
        "这个文件仅作格式参考，真实更新请先从后台导出现有商品后再修改。",
        "https://images.unsplash.com/photo-1514733670139-4d87a19b179d?w=600",
        "https://images.unsplash.com/photo-1514733670139-4d87a19b179d?w=600",
        "更新已有商品时，必须保留同一商品全部 SKU 行。",
        "https://images.unsplash.com/photo-1514733670139-4d87a19b179d?w=900",
    ]
    rows = []
    for sku_code, spec_value, price, market_price, stock in [
        ("XYS-100G", "100g", "88.00", "108.00", "20"),
        ("XYS-200G", "200g", "158.00", "188.00", "12"),
    ]:
        rows.append(common + [
            sku_code,
            "净含量",
            spec_value,
            price,
            market_price,
            stock,
            "上架",
            "90",
            "https://images.unsplash.com/photo-1514733670139-4d87a19b179d?w=600",
            "",
        ])
    return rows


def upsert_rows() -> list[list[str]]:
    common = [
        "SPU-10002",
        "示例：给老商品补新规格",
        "实物",
        "",
        "轻养零食",
        "示例 老商品 补规格",
        "这个文件仅作格式参考，真实补规格请先导出现有商品再新增一行 SKU。",
        "https://images.unsplash.com/photo-1595855759920-86582396756a?w=600",
        "https://images.unsplash.com/photo-1595855759920-86582396756a?w=600",
        "已有规格和新增规格都要一起保留在导入文件中。",
        "https://images.unsplash.com/photo-1595855759920-86582396756a?w=900",
    ]
    rows = []
    for sku_code, spec_value, price, market_price, stock in [
        ("HZMW-30", "30丸", "42.00", "56.00", "60"),
        ("HZMW-60", "60丸", "79.00", "99.00", "35"),
    ]:
        rows.append(common + [
            sku_code,
            "规格",
            spec_value,
            price,
            market_price,
            stock,
            "上架",
            "80",
            "https://images.unsplash.com/photo-1595855759920-86582396756a?w=600",
            "",
        ])
    return rows


def save_workbook(path: Path, title: str, rows: list[list[str]]) -> None:
    workbook = build_workbook(title, rows)
    workbook.save(path)


def main() -> None:
    root = Path(__file__).resolve().parent.parent
    output_dir = root / "docs" / "delivery" / "product-template-pack"
    output_dir.mkdir(parents=True, exist_ok=True)

    save_workbook(output_dir / "01-商品导入空白模板.xlsx", "商品导入空白模板", [])
    save_workbook(output_dir / "02-新增商品示例-单规格.xlsx", "新增商品示例 - 单规格", single_spec_rows())
    save_workbook(output_dir / "03-新增商品示例-多规格.xlsx", "新增商品示例 - 多规格", multi_spec_rows())
    save_workbook(output_dir / "04-更新商品示例-改价改库存.xlsx", "更新商品示例 - 仅作格式参考", update_rows())
    save_workbook(output_dir / "05-补充规格示例-新增SKU.xlsx", "补充规格示例 - 仅作格式参考", upsert_rows())

    readme_path = output_dir / "README-商品模板包说明.txt"
    readme_path.write_text(README_TEXT, encoding="utf-8")

    zip_path = output_dir / "药食同源-v1.0-商品Excel模板包.zip"
    with ZipFile(zip_path, "w", compression=ZIP_DEFLATED) as archive:
        for file_path in sorted(output_dir.iterdir()):
            if file_path.name == zip_path.name:
                continue
            archive.write(file_path, arcname=file_path.name)

    print(output_dir)
    print(zip_path)


if __name__ == "__main__":
    main()
